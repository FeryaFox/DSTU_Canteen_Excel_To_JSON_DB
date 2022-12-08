from openpyxl import load_workbook
import os
import json

file_list = []

for i in os.listdir():
    if "xlsx" in i.lower() or "xls" in i.lower():
        file_list.append(i)

while True:
    print()
    j = 1
    for i in file_list:
        print(f"{j} - {i}")
        j += 1
    chose_file_id = input("Выберите файл, который нужно конвертировать: ")

    if chose_file_id.isdigit():
        break

chose_file = file_list[int(chose_file_id)-1]
print(chose_file)

xlsx_file = load_workbook(chose_file)

final_dict = {"canteen_list": [], "dishes": []}
canteen_id = 0

for sheetname in xlsx_file.sheetnames:
    final_dict["canteen_list"].append({"id": canteen_id, "name": sheetname, "description": "", "place": "", "photo": []})
    dishes_id = 0
    for row in xlsx_file[sheetname].iter_rows():
        if row[0].value == "Name":
            continue
        print(row[0].value)
        final_dict["dishes"].append(
            {
                "id": dishes_id,
                "canteen_id": canteen_id,
                "name": str(row[0].value),
                "category": str(row[1].value),
                "price": float(row[2].value)
            }
        )
        dishes_id += 1
    canteen_id += 1

print(final_dict)

with open("canteen_data.json", "w", encoding='utf-8') as f:
    json.dump(final_dict, f, ensure_ascii=False, indent=4)
